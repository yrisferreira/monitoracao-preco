import os
import time
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openai import OpenAI
from config import *
import openpyxl

# Carrega variáveis de ambiente
load_dotenv()

class MonitorPrecos:
    def __init__(self):
        self.driver = self._configurar_driver()
        self.resultados = []
        try:
            self.client = OpenAI()
            self.client.api_key = OPENAI_API_KEY
        except Exception as e:
            print(f"Erro ao inicializar cliente OpenAI: {e}")
            self.client = None
        
    def _configurar_driver(self):
        """Configura o driver do Selenium"""
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument(f"user-agent={USER_AGENT}")
        
        # Configuração específica para macOS
        service = Service()
        return webdriver.Chrome(service=service, options=chrome_options)

    def _obter_preco_sephora(self, url):
        """Obtém preço do produto na Sephora"""
        try:
            self.driver.get(url)
            time.sleep(TEMPO_ESPERA)
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Tenta diferentes seletores comuns para preços
            preco_element = (
                soup.find('span', {'class': 'price'}) or
                soup.find('span', {'class': 'product-price'}) or
                soup.find('div', {'class': 'price'}) or
                soup.find('span', {'data-testid': 'price-value'}) or
                soup.find('span', {'class': 'price-value'})
            )
            
            if preco_element and preco_element.text:
                preco_texto = preco_element.text.strip()
                # Remove caracteres não numéricos exceto ponto e vírgula
                preco_texto = ''.join(c for c in preco_texto if c.isdigit() or c in '.,')
                return float(preco_texto.replace('.', '').replace(',', '.'))
            else:
                print(f"Elemento de preço não encontrado na Sephora para URL: {url}")
                return None
                
        except Exception as e:
            print(f"Erro ao obter preço da Sephora: {e}")
            return None

    def _obter_preco_boticario(self, url):
        """Obtém preço do produto no Boticário"""
        try:
            self.driver.get(url)
            time.sleep(TEMPO_ESPERA)
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Tenta diferentes seletores comuns para preços
            preco_element = (
                soup.find('span', {'class': 'price'}) or
                soup.find('span', {'class': 'product-price'}) or
                soup.find('div', {'class': 'price'}) or
                soup.find('span', {'data-testid': 'price-value'}) or
                soup.find('span', {'class': 'price-value'})
            )
            
            if preco_element and preco_element.text:
                preco_texto = preco_element.text.strip()
                # Remove caracteres não numéricos exceto ponto e vírgula
                preco_texto = ''.join(c for c in preco_texto if c.isdigit() or c in '.,')
                return float(preco_texto.replace('.', '').replace(',', '.'))
            else:
                print(f"Elemento de preço não encontrado no Boticário para URL: {url}")
                return None
                
        except Exception as e:
            print(f"Erro ao obter preço do Boticário: {e}")
            return None

    def _obter_preco_amazon(self, url):
        """Obtém preço do produto na Amazon"""
        try:
            self.driver.get(url)
            time.sleep(TEMPO_ESPERA)
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Tenta diferentes seletores comuns para preços na Amazon
            preco_element = (
                soup.find('span', {'class': 'a-price-whole'}) or
                soup.find('span', {'class': 'a-offscreen'}) or
                soup.find('span', {'class': 'a-price'}) or
                soup.find('span', {'data-a-color': 'price'}) or
                soup.find('span', {'class': 'a-price-range'})
            )
            
            if preco_element and preco_element.text:
                preco_texto = preco_element.text.strip()
                # Remove caracteres não numéricos exceto ponto e vírgula
                preco_texto = ''.join(c for c in preco_texto if c.isdigit() or c in '.,')
                return float(preco_texto.replace('.', '').replace(',', '.'))
            else:
                print(f"Elemento de preço não encontrado na Amazon para URL: {url}")
                return None
                
        except Exception as e:
            print(f"Erro ao obter preço da Amazon: {e}")
            return None

    def _obter_preco_paraguai(self, url):
        """Obtém preço do produto no site de compras no Paraguai"""
        try:
            self.driver.get(url)
            time.sleep(TEMPO_ESPERA)
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Tenta diferentes seletores comuns para preços
            preco_element = (
                soup.find('span', {'class': 'price'}) or
                soup.find('span', {'class': 'product-price'}) or
                soup.find('div', {'class': 'price'}) or
                soup.find('span', {'data-testid': 'price-value'}) or
                soup.find('span', {'class': 'price-value'})
            )
            
            if preco_element and preco_element.text:
                preco_texto = preco_element.text.strip()
                # Remove caracteres não numéricos exceto ponto e vírgula
                preco_texto = ''.join(c for c in preco_texto if c.isdigit() or c in '.,')
                return float(preco_texto.replace('.', '').replace(',', '.'))
            else:
                print(f"Elemento de preço não encontrado no site de compras no Paraguai para URL: {url}")
                return None
                
        except Exception as e:
            print(f"Erro ao obter preço do site de compras no Paraguai: {e}")
            return None

    def _obter_preco_apple(self, url):
        """Obtém preço do produto na Apple Store"""
        try:
            self.driver.get(url)
            time.sleep(TEMPO_ESPERA)
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            
            # Tenta diferentes seletores comuns para preços na Apple Store
            preco_element = (
                soup.find('span', {'class': 'price'}) or
                soup.find('span', {'class': 'product-price'}) or
                soup.find('div', {'class': 'price'}) or
                soup.find('span', {'data-testid': 'price-value'}) or
                soup.find('span', {'class': 'price-value'}) or
                soup.find('span', {'class': 'current-price'})
            )
            
            if preco_element and preco_element.text:
                preco_texto = preco_element.text.strip()
                # Remove caracteres não numéricos exceto ponto e vírgula
                preco_texto = ''.join(c for c in preco_texto if c.isdigit() or c in '.,')
                return float(preco_texto.replace('.', '').replace(',', '.'))
            else:
                print(f"Elemento de preço não encontrado na Apple Store para URL: {url}")
                return None
                
        except Exception as e:
            print(f"Erro ao obter preço da Apple Store: {e}")
            return None

    def monitorar_precos(self):
        """Monitora preços de todos os produtos configurados"""
        for produto in PRODUTOS:
            resultado = {
                'produto': produto['nome'],
                'data': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'preco_alvo': produto['preco_alvo']
            }

            # Obtém preços de cada loja
            for loja, url in produto['urls'].items():
                if loja == 'sephora':
                    preco = self._obter_preco_sephora(url)
                elif loja == 'boticario':
                    preco = self._obter_preco_boticario(url)
                elif loja == 'amazon':
                    preco = self._obter_preco_amazon(url)
                elif loja == 'paraguai':
                    preco = self._obter_preco_paraguai(url)
                elif loja == 'apple':
                    preco = self._obter_preco_apple(url)
                
                resultado[f'preco_{loja}'] = preco

            self.resultados.append(resultado)

    def gerar_excel(self):
        """Gera planilha Excel com os resultados"""
        if not os.path.exists(PASTA_DADOS):
            os.makedirs(PASTA_DADOS)
        
        df = pd.DataFrame(self.resultados)
        
        # Formata as colunas de preço
        colunas_preco = [col for col in df.columns if col.startswith('preco_')]
        for col in colunas_preco:
            df[col] = df[col].apply(lambda x: f'R$ {x:.2f}' if pd.notnull(x) else 'Não disponível')
        
        # Formata a coluna de preço alvo
        df['preco_alvo'] = df['preco_alvo'].apply(lambda x: f'R$ {x:.2f}')
        
        # Renomeia as colunas para melhor legibilidade
        df = df.rename(columns={
            'produto': 'Produto',
            'data': 'Data/Hora',
            'preco_alvo': 'Preço Alvo',
            'preco_sephora': 'Preço Sephora',
            'preco_boticario': 'Preço Boticário',
            'preco_amazon': 'Preço Amazon',
            'preco_paraguai': 'Preço Paraguai',
            'preco_apple': 'Preço Apple'
        })
        
        # Reordena as colunas
        colunas_ordenadas = ['Produto', 'Data/Hora', 'Preço Alvo', 'Preço Apple', 'Preço Sephora', 'Preço Boticário', 'Preço Amazon', 'Preço Paraguai']
        df = df[colunas_ordenadas]
        
        # Salva o arquivo Excel
        caminho_arquivo = os.path.join(PASTA_DADOS, NOME_ARQUIVO_EXCEL)
        
        # Cria um writer do Excel
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Preços')
            
            # Obtém a planilha
            worksheet = writer.sheets['Preços']
            
            # Ajusta a largura das colunas
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2
            
            # Formata o cabeçalho
            for cell in worksheet[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Formata as células de preço
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    
                    # Destaca preços abaixo do alvo em verde
                    if cell.column > 3 and cell.value != 'Não disponível':  # Colunas de preço
                        preco = float(cell.value.replace('R$ ', '').replace('.', '').replace(',', '.'))
                        preco_alvo = float(worksheet.cell(row=cell.row, column=3).value.replace('R$ ', '').replace('.', '').replace(',', '.'))
                        if preco < preco_alvo:
                            cell.fill = openpyxl.styles.PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        print(f"Planilha gerada: {caminho_arquivo}")

    def enviar_alerta_email(self):
        """Envia e-mail de alerta quando preço está abaixo do alvo"""
        for resultado in self.resultados:
            precos = [v for k, v in resultado.items() if k.startswith('preco_') and v is not None]
            if precos and min(precos) < resultado['preco_alvo']:
                self._enviar_email(resultado)

    def _enviar_email(self, resultado):
        """Envia e-mail de alerta"""
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['email_from']
        msg['To'] = EMAIL_CONFIG['email_to']
        msg['Subject'] = f"Alerta de Preço: {resultado['produto']}"

        corpo = f"""
        Produto: {resultado['produto']}
        Preço alvo: R$ {resultado['preco_alvo']:.2f}
        
        Preços encontrados:
        """
        for k, v in resultado.items():
            if k.startswith('preco_'):
                loja = k.replace('preco_', '').title()
                if v is not None:
                    corpo += f"\n{loja}: R$ {v:.2f}"

        msg.attach(MIMEText(corpo, 'plain'))

        try:
            server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
            server.starttls()
            server.login(EMAIL_CONFIG['email_from'], os.getenv('EMAIL_PASSWORD'))
            server.send_message(msg)
            server.quit()
            print(f"E-mail de alerta enviado para {resultado['produto']}")
        except Exception as e:
            print(f"Erro ao enviar e-mail: {e}")

    def analisar_tendencias(self):
        """Usa IA para analisar tendências de preços"""
        if not self.resultados or not self.client:
            return

        # Prepara dados para análise
        dados = pd.DataFrame(self.resultados)
        prompt = f"""
        Analise os seguintes dados de preços e forneça insights sobre:
        1. Melhor momento para compra
        2. Tendência de preços
        3. Recomendações

        Dados:
        {dados.to_string()}
        """

        try:
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "Você é um especialista em análise de preços e tendências de mercado."},
                    {"role": "user", "content": prompt}
                ]
            )
            print("\nAnálise de Tendências:")
            print(response.choices[0].message.content)
        except Exception as e:
            print(f"Erro na análise de tendências: {e}")

    def fechar(self):
        """Fecha o driver do Selenium"""
        self.driver.quit()

def main():
    monitor = MonitorPrecos()
    try:
        monitor.monitorar_precos()
        monitor.gerar_excel()
        monitor.enviar_alerta_email()
        monitor.analisar_tendencias()
    finally:
        monitor.fechar()

if __name__ == "__main__":
    main() 